import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime
from faker import Faker
from datetime import datetime
from datetime import timedelta

fake = Faker('fr_CH')  # Utilisation de la langue française et des formats spécifiques à la Suisse

def generate_data(num_rows):
    data = []
    for _ in range(num_rows):
        nom = fake.last_name()
        prenom = fake.first_name()
        dateN = fake.date_of_birth(minimum_age=18, maximum_age=90).strftime('%Y-%m-%d')
        adresse = fake.street_address()
        canton = fake.country()
        assurance = fake.company()
        
        # AssureDep est calculé comme 1 jour après DateN
        dateN_datetime = datetime.strptime(dateN, '%Y-%m-%d')
        assure_dep = (dateN_datetime + timedelta(days=1)).strftime('%Y-%m-%d')
        
        permis_id = fake.random_number(digits=3)

        data.append((str(permis_id).zfill(3), nom, prenom, dateN, adresse, str(canton), assurance, assure_dep))
    
    return data



def create_excel_file(data):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Entête
    headers = ['permisID', 'nom', 'prenom', 'dateN', 'adresse', 'canton', 'assurance', 'assureDep']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Remplissage des données
    unique_permis_ids = set()
    row_idx = 2
    for row in data:
        permis_id, nom, prenom, dateN, adresse, canton, assurance, assure_dep = row
        
        # Vérification des conditions
        if (permis_id in unique_permis_ids or
            not permis_id.isalnum() or
            datetime.strptime(dateN, '%Y-%m-%d') > datetime.strptime(assure_dep, '%Y-%m-%d') or
            not canton.isalpha() or
            '' in [nom, prenom, adresse, assurance]):  # Ignorer si une valeur est une chaîne vide
            continue
        
        unique_permis_ids.add(permis_id)

        # Insertion des données dans la feuille Excel
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        
        row_idx += 1

    # Sauvegarde du fichier Excel
    wb.save("donnees.xlsx")

# Exemple de données
# donnees = [
#     ('001', 'Doe', 'John', '1989-12-31', '123 Rue de Paris', 'Geneva', 'AXA', '1990-01-01'),
#     ('002', 'Smith', 'Alice', '1985-05-10', '456 Avenue de Lyon', 'Zurich', 'Allianz', '1999-05-15'),
#     ('003', 'Dupont', 'Pierre', '1978-09-19', '789 Boulevard de Marseille', 'Lausanne', 'Generali', '1978-09-20'),
#     ('002', 'Dupuis', 'Jean', '1980-03-24 ', '987 Rue de Bordeaux', 'Lugano', 'AXA', '1980-03-25'),  # Doublon
#     ('004', 'Smith', '', '1985-05-10', '456 Avenue de Lyon', 'Zurich', 'Allianz', '1999-05-15'), # chaine vide
#     ('005', 'Smith', 'jean', '1999-05-10', '456 Avenue de Lyon', 'Zurich', 'Allianz', '1988-05-15'), # date 
#     ('006', 'Smi,th,', 'jean', '1999-05-10', '456 Avenue de Lyon', 'Zurich', 'Allianz', '1988-05-15'), # Pas de caracteres speciaux
#     ('007', 'Smith test', 'jean', '1999-05-10', '456 Avenue de Lyon', 'ZR', 'Allianz', '1988-05-15') # Doit respecter le format de pays
# ]

# Générer des données au format spécifié
donnees = generate_data(10)

# Afficher les données
print(donnees)

create_excel_file(donnees)